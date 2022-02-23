# openExcel.py
# short script using openpyxl to read and print all cell data

import openpyxl as xl


def fuckYouHarder(x, y):
    print("uuuuuuuuuuuuuuuuuuuuuuuuuuuuuuuugh")
    return x-y


def fuckYou():
    print("POOOOOOOOOOOOOOOP")
    x = 1 + 1
    y = 2 + 11
    print(str(x+y))
    print(fuckYouHarder(x, y))


if __name__ == '__main__':
    # read workbook
    wb = xl.load_workbook('..\Downloads\Checklist.xlsx')
    print(wb.sheetnames)
    s = wb.active
    if s == None: quit()    # failsafe

    fuckYou()

    # just want to print all cells in first sheet
    for row in range(1, s.max_row + 1):
        row_val = str(s.cell(row, 1).value)
        if row_val != "None":
            for col in range(1, s.max_column + 1):
                val = str(s.cell(row, col).value)
                if val != "None":
                    print(val + " | ", end='')
            print()
